"""Reportes: Excel consolidado del backlog y métricas de ciclo/efectividad."""
from pathlib import Path
from io import BytesIO
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .config import ICON_OK, ICON_ERROR, ICON_WARNING, ICON_SUCCESS, ICON_FAIL
from .analysis import get_estado_code, ESTADO_LISTO, ESTADO_ERROR, ESTADO_INCOMPLETO, ESTADO_SIN_METADATA


def calcular_eficiencia_por_ciclo(created_date: str, downloaded_at: str, changed_date: str, estado_ado: str) -> tuple:
    """Calcula días de ciclo SOLO si estado_ado == 'Closed'
    Retorna (días_creación_a_cierre, días_descarga_a_cierre) o (0, 0) si aún está abierto"""
    if estado_ado != "Closed":
        return 0, 0

    try:
        fecha_creacion = datetime.fromisoformat(created_date.replace("Z", "+00:00"))
        fecha_descarga = datetime.fromisoformat(downloaded_at.replace("Z", "+00:00")) if downloaded_at else datetime.now()
        fecha_cierre   = datetime.fromisoformat(changed_date.replace("Z", "+00:00")) if changed_date else datetime.now()
        return (fecha_cierre - fecha_creacion).days, (fecha_cierre - fecha_descarga).days
    except:
        return 0, 0


def generar_excel_consolidado(resultados: list, guardar_en_carpeta: Path = None) -> bytes:
    """Genera Excel con Sprint, Estado ADO real, fechas de ciclo y métricas (solo para Closed)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"

    #  HEADERS
    headers = ["ID", "Título", "Tipo", "Sprint", "Estado ADO",
               "Fecha Creación", "Fecha Descarga", "Fecha Cierre",
               "TA", "AID", "UDZ", "RNF", "Días Creación→Cierre", "Días Descarga→Cierre",
               "Aprobado por", "Fecha aprobación"]

    ws.append(headers)

    # Estilos
    header_fill = PatternFill(start_color="FDDA24", end_color="FDDA24", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Aplicar estilos a headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    #  DATOS
    row = 2
    for r in sorted(resultados, key=lambda x: x.get("downloaded_at", ""), reverse=True):
        hu_id = r.get("hu_id", "")
        title = r.get("hu_title", "")[:50]
        tipo = r.get("tipo_cambio", "")
        sprint = r.get("sprint", "?")
        estado_ado = r.get("estado_ado", "New")

        # Fechas reales de ADO
        created_date = r.get("created_date", "")
        downloaded_at = r.get("downloaded_at", "")
        changed_date = r.get("changed_date", "")

        # Parsear fechas para mostrar
        fecha_creacion = created_date[:10] if created_date else "?"
        fecha_descarga = downloaded_at[:10] if downloaded_at else "?"

        # Fecha de cierre: solo si está Closed
        if estado_ado == "Closed":
            fecha_cierre = changed_date[:10] if changed_date else "?"
        else:
            fecha_cierre = "-"  # Sin cerrar aún

        # Calcular ciclos (solo si está Closed)
        dias_creacion_cierre, dias_descarga_cierre = calcular_eficiencia_por_ciclo(
            created_date, downloaded_at, changed_date, estado_ado
        )

        # Si no está Closed, mostrar guiones en días
        dias_creacion_str = dias_creacion_cierre if dias_creacion_cierre > 0 else "-"
        dias_descarga_str = dias_descarga_cierre if dias_descarga_cierre > 0 else "-"

        # Archivos
        arcs = r.get("archivos", {})
        ta = ICON_SUCCESS if "NO" not in arcs.get("TA", "") else ICON_FAIL
        aid = ICON_SUCCESS if "NO" not in arcs.get("AID", "") else ICON_FAIL
        udz = ICON_SUCCESS if "NO" not in arcs.get("UDZ", "") else ICON_FAIL
        rnf = ICON_SUCCESS if r.get("rnf_path") else ICON_FAIL

        # Trazabilidad de aprobación para PDN
        aprobado_por = r.get("aprobado_por") or "-"
        aprobado_en_raw = r.get("aprobado_en", "")
        aprobado_en = aprobado_en_raw[:16].replace("T", " ") if aprobado_en_raw else "-"

        # Agregar fila con nuevas fechas (sin Ambiente)
        ws.append([
            hu_id, title, tipo, sprint, estado_ado,
            fecha_creacion, fecha_descarga, fecha_cierre,
            ta, aid, udz, rnf,
            dias_creacion_str, dias_descarga_str,
            aprobado_por, aprobado_en
        ])

        # Aplicar estilos a fila
        for cell in ws[row]:
            cell.border = border
            if ICON_SUCCESS in str(cell.value):
                cell.font = Font(color="008000")
            elif ICON_FAIL in str(cell.value):
                cell.font = Font(color="FF0000")

        row += 1

    #  AJUSTAR ANCHO COLUMNAS
    ws.column_dimensions["A"].width = 12   # ID
    ws.column_dimensions["B"].width = 35   # Título
    ws.column_dimensions["C"].width = 15   # Tipo
    ws.column_dimensions["D"].width = 18   # Sprint
    ws.column_dimensions["E"].width = 12   # Estado ADO
    ws.column_dimensions["F"].width = 14   # Fecha Creación
    ws.column_dimensions["G"].width = 14   # Fecha Descarga
    ws.column_dimensions["H"].width = 14   # Fecha Cierre
    ws.column_dimensions["I"].width = 5    # TA
    ws.column_dimensions["J"].width = 5    # AID
    ws.column_dimensions["K"].width = 5    # UDZ
    ws.column_dimensions["L"].width = 5    # RNF
    ws.column_dimensions["M"].width = 20   # Días Creación→Cierre
    ws.column_dimensions["N"].width = 20   # Días Descarga→Cierre
    ws.column_dimensions["O"].width = 18   # Aprobado por
    ws.column_dimensions["P"].width = 18   # Fecha aprobación

    #  HOJA: EFECTIVIDAD
    ws_ef = wb.create_sheet("Efectividad")

    # Agrupar resultados por sprint
    sprints_data = defaultdict(list)
    for r in resultados:
        sprints_data[r.get("sprint", "Sin Sprint")].append(r)

    ef_headers = [
        "Sprint", "Total HU", f"{ICON_OK} Listos", f"{ICON_ERROR} Con Errores", f"{ICON_WARNING} Incompletos",
        "% éxito", "Errores en TA", "Errores en AID", "Errores en UDZ",
        "DESPLIEGUE", "MODIFICACIÓN", "Fecha Análisis"
    ]
    ws_ef.append(ef_headers)

    for cell in ws_ef[1]:
        cell.fill = PatternFill(start_color="2C2A29", end_color="2C2A29", fill_type="solid")
        cell.font = Font(bold=True, color="FDDA24")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ef_row = 2
    for sprint_name, hu_list in sorted(sprints_data.items()):
        total      = len(hu_list)
        listos     = sum(1 for r in hu_list if get_estado_code(r) == ESTADO_LISTO)
        errores    = sum(1 for r in hu_list if get_estado_code(r) == ESTADO_ERROR)
        incompl    = sum(1 for r in hu_list if get_estado_code(r) in (ESTADO_INCOMPLETO, ESTADO_SIN_METADATA))
        pct        = f"{round(listos / total * 100)}%" if total else "0%"

        # Errores por componente: cualquier validación que depende de ese archivo y falló
        err_ta = err_aid = err_udz = 0
        for r in hu_list:
            v = r.get("validaciones", {})
            arcs = r.get("archivos", {})
            if "NO" not in arcs.get("TA", "NO"):
                if not v.get("kafka", {}).get("ok", True) or not v.get("coherencia", {}).get("ok", True):
                    err_ta += 1
            if "NO" not in arcs.get("AID", "NO"):
                if not v.get("s3_path", {}).get("ok", True) or not v.get("last_step", {}).get("ok", True) or not v.get("out_zone_copiar", {}).get("out_zone_ok", True):
                    err_aid += 1
            if "NO" not in arcs.get("UDZ", "NO"):
                if not v.get("s3_path", {}).get("ok", True) or not v.get("workflow_vs_id", {}).get("ok", True):
                    err_udz += 1

        desp = sum(1 for r in hu_list if "DESP" in r.get("tipo_cambio", "").upper())
        modi = sum(1 for r in hu_list if "MODI" in r.get("tipo_cambio", "").upper())

        # Fecha del análisis más reciente del sprint
        fechas = [r.get("downloaded_at", "")[:10] for r in hu_list if r.get("downloaded_at")]
        fecha_analisis = max(fechas) if fechas else datetime.now().strftime("%Y-%m-%d")

        ws_ef.append([sprint_name, total, listos, errores, incompl, pct,
                       err_ta, err_aid, err_udz, desp, modi, fecha_analisis])

        # Color de fila según % éxito
        fill_row = PatternFill(
            start_color="D1FAE5" if listos == total else ("FEF3C7" if errores < total // 2 else "FEE2E2"),
            end_color  ="D1FAE5" if listos == total else ("FEF3C7" if errores < total // 2 else "FEE2E2"),
            fill_type="solid"
        )
        for cell in ws_ef[ef_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill_row
        ef_row += 1

    # Anchos
    for col, w in zip("ABCDEFGHIJKL", [22, 10, 12, 14, 14, 10, 12, 12, 12, 14, 14, 16]):
        ws_ef.column_dimensions[col].width = w

    # Guardar a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Si se proporciona una carpeta, guardar también allí con nombre estándar
    if guardar_en_carpeta:
        guardar_en_carpeta.mkdir(parents=True, exist_ok=True)
        archivo_path = guardar_en_carpeta / "Consolidado_Backlog.xlsx"
        wb.save(str(archivo_path))

    return output.getvalue()
